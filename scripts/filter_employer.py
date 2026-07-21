#!/usr/bin/env python3
"""Filter DOL LCA disclosure xlsx files for one employer (regex over name/DBA).

Usage:
    filter_employer.py "<regex>" out.csv data/LCA_FY2025.xlsx [more.xlsx ...]

Streams each workbook read-only (low memory). Matches EMPLOYER_NAME or
TRADE_NAME_DBA case-insensitively against the regex. Writes a slim CSV of the
columns that matter for role-level analysis, plus a FISCAL_YEAR tag inferred
from the source filename.
"""
import csv
import re
import sys
from pathlib import Path

import openpyxl

KEEP = [
    "CASE_NUMBER", "CASE_STATUS", "RECEIVED_DATE", "DECISION_DATE", "VISA_CLASS",
    "JOB_TITLE", "SOC_TITLE", "FULL_TIME_POSITION", "BEGIN_DATE", "END_DATE",
    "TOTAL_WORKER_POSITIONS", "NEW_EMPLOYMENT", "CONTINUED_EMPLOYMENT",
    "CHANGE_EMPLOYER", "AMENDED_PETITION", "EMPLOYER_NAME", "TRADE_NAME_DBA",
    "EMPLOYER_CITY", "EMPLOYER_STATE", "NAICS_CODE",
    "SECONDARY_ENTITY", "SECONDARY_ENTITY_BUSINESS_NAME",
    "WORKSITE_CITY", "WORKSITE_COUNTY", "WORKSITE_STATE", "WORKSITE_POSTAL_CODE",
    "WAGE_RATE_OF_PAY_FROM", "WAGE_RATE_OF_PAY_TO", "WAGE_UNIT_OF_PAY",
    "PREVAILING_WAGE", "PW_UNIT_OF_PAY", "PW_WAGE_LEVEL", "PW_OTHER_SOURCE",
    "H_1B_DEPENDENT", "WILLFUL_VIOLATOR",
]


def fy_from_name(path: str) -> str:
    m = re.search(r"FY(\d{4})", path)
    return m.group(1) if m else "?"


def main():
    if len(sys.argv) < 4:
        sys.exit(__doc__)
    pattern = re.compile(sys.argv[1], re.I)
    out_path = sys.argv[2]
    sources = sys.argv[3:]

    matched = 0
    with open(out_path, "w", newline="") as fh:
        writer = csv.writer(fh)
        writer.writerow(["FISCAL_YEAR"] + KEEP)
        for src in sources:
            fy = fy_from_name(src)
            wb = openpyxl.load_workbook(src, read_only=True)
            ws = wb[wb.sheetnames[0]]
            it = ws.iter_rows(values_only=True)
            header = list(next(it))
            idx = {name: header.index(name) for name in header}
            name_i = idx["EMPLOYER_NAME"]
            dba_i = idx["TRADE_NAME_DBA"]
            keep_i = [idx[c] for c in KEEP]
            n = 0
            for row in it:
                n += 1
                name = str(row[name_i] or "")
                dba = str(row[dba_i] or "")
                if pattern.search(name) or pattern.search(dba):
                    writer.writerow([fy] + [row[i] for i in keep_i])
                    matched += 1
            wb.close()
            print(f"  scanned {n:,} rows in {Path(src).name}", file=sys.stderr)
    print(f"matched {matched} rows -> {out_path}", file=sys.stderr)


if __name__ == "__main__":
    main()

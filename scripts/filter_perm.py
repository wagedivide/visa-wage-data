#!/usr/bin/env python3
"""Filter DOL PERM (green-card labor certification) disclosure files by employer.

Usage:
    filter_perm.py "<regex>" out.csv file1.xlsx [file2.xlsx ...]

PERM is the permanent-residency step: the employer must "test the U.S. labor
market" and attest no qualified American is available. This script pulls the
role, wage, worksite, recruitment tactics, and the sham-recruitment red-flag
fields (foreign-language requirement, requirements exceeding the norm,
layoff-in-area) for any employer. Schema differs between the FY2024 old form and
the FLAG new form, so columns are matched by name if present, else left blank.
"""
import csv
import re
import sys
from pathlib import Path

import openpyxl

# canonical name -> list of possible header spellings across form versions
WANT = {
    "CASE_STATUS": ["CASE_STATUS"],
    "DECISION_DATE": ["DECISION_DATE"],
    "EMPLOYER": ["EMP_BUSINESS_NAME", "EMPLOYER_NAME"],
    "EMP_STATE": ["EMP_STATE", "EMPLOYER_STATE"],
    "EMP_PAYROLL": ["EMP_NUM_PAYROLL", "EMPLOYER_NUM_EMPLOYEES"],
    "SOC_TITLE": ["PWD_SOC_TITLE", "PW_SOC_TITLE"],
    "JOB_TITLE": ["JOB_TITLE", "JOB_INFO_JOB_TITLE"],
    "WAGE_FROM": ["JOB_OPP_WAGE_FROM", "WAGE_OFFER_FROM_9089", "WAGE_OFFERED_FROM"],
    "WAGE_TO": ["JOB_OPP_WAGE_TO", "WAGE_OFFER_TO_9089"],
    "WAGE_PER": ["JOB_OPP_WAGE_PER", "WAGE_OFFER_UNIT_OF_PAY_9089"],
    "WORKSITE_CITY": ["PRIMARY_WORKSITE_CITY", "JOB_INFO_WORK_CITY"],
    "WORKSITE_STATE": ["PRIMARY_WORKSITE_STATE", "JOB_INFO_WORK_STATE"],
    "WORKSITE_ZIP": ["PRIMARY_WORKSITE_POSTAL_CODE", "JOB_INFO_WORK_POSTAL_CODE"],
    "FOREIGN_LANG_REQ": ["OTHER_REQ_JOB_FOREIGN_LANGUAGE", "JOB_INFO_FOREIGN_LANG_REQUIRED"],
    "REQ_EXCEEDS_NORM": ["OTHER_REQ_JOB_REQ_EXCEED", "JOB_INFO_EXPERIENCE_REQD"],
    "LAYOFF_IN_AREA": ["OTHER_REQ_EMP_LAYOFF", "RECR_INFO_EMPLOYER_LAYOFF"],
    "NEWSPAPER_SUNDAY": ["RECR_INFO_IS_NEWSPAPER_SUNDAY"],
    "NEWSPAPER_NAME": ["RECR_INFO_NEWSPAPER_NAME"],
    "AD_TYPE": ["RECR_INFO_RECRUIT_AD_TYPE"],
    "LAW_FIRM": ["ATTY_AG_LAW_FIRM_NAME", "AGENT_ATTORNEY_LAW_FIRM_NAME"],
}


def fy(path):
    m = re.search(r"FY(\d{4})", path)
    return m.group(1) if m else "?"


def main():
    if len(sys.argv) < 4:
        sys.exit(__doc__)
    pat = re.compile(sys.argv[1], re.I)
    out_path = sys.argv[2]
    sources = sys.argv[3:]

    matched = 0
    with open(out_path, "w", newline="") as fh:
        w = csv.writer(fh)
        w.writerow(["FISCAL_YEAR"] + list(WANT.keys()))
        for src in sources:
            y = fy(src)
            wb = openpyxl.load_workbook(src, read_only=True)
            ws = wb[wb.sheetnames[0]]
            it = ws.iter_rows(values_only=True)
            header = [str(h) if h is not None else "" for h in next(it)]
            hidx = {h: i for i, h in enumerate(header)}
            # resolve each canonical col to an index (or None)
            col_i = {}
            for canon, spellings in WANT.items():
                col_i[canon] = next((hidx[s] for s in spellings if s in hidx), None)
            emp_i = col_i["EMPLOYER"]
            if emp_i is None:
                print(f"  !! no employer column in {Path(src).name}; headers start: {header[:8]}", file=sys.stderr)
                wb.close()
                continue
            n = 0
            for row in it:
                n += 1
                if pat.search(str(row[emp_i] or "")):
                    w.writerow([y] + [(row[col_i[c]] if col_i[c] is not None else "") for c in WANT])
                    matched += 1
            wb.close()
            print(f"  scanned {n:,} in {Path(src).name}", file=sys.stderr)
    print(f"matched {matched} -> {out_path}", file=sys.stderr)


if __name__ == "__main__":
    main()

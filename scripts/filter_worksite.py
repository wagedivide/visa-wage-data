#!/usr/bin/env python3
"""Filter DOL LCA disclosure files by WORKSITE state + city list.

Usage:
    filter_worksite.py <state> "<city1,city2,...>" out.csv file1.xlsx [file2.xlsx ...]

Matches WORKSITE_STATE exactly (case-insensitive) and WORKSITE_CITY against any
of the given cities (case-insensitive, substring). Writes a slim CSV of the
role-level columns + FISCAL_YEAR. Certified-only rows are kept (a seated role
required a certified LCA).
"""
import csv
import re
import sys
from pathlib import Path

import openpyxl

KEEP = [
    "CASE_NUMBER", "CASE_STATUS", "DECISION_DATE", "VISA_CLASS",
    "JOB_TITLE", "SOC_TITLE", "BEGIN_DATE", "END_DATE",
    "TOTAL_WORKER_POSITIONS", "NEW_EMPLOYMENT", "CONTINUED_EMPLOYMENT",
    "CHANGE_EMPLOYER", "EMPLOYER_NAME", "TRADE_NAME_DBA",
    "EMPLOYER_CITY", "EMPLOYER_STATE", "NAICS_CODE",
    "SECONDARY_ENTITY", "SECONDARY_ENTITY_BUSINESS_NAME",
    "WORKSITE_CITY", "WORKSITE_COUNTY", "WORKSITE_STATE", "WORKSITE_POSTAL_CODE",
    "WAGE_RATE_OF_PAY_FROM", "WAGE_RATE_OF_PAY_TO", "WAGE_UNIT_OF_PAY",
    "PREVAILING_WAGE", "PW_UNIT_OF_PAY", "PW_WAGE_LEVEL",
    "H_1B_DEPENDENT", "WILLFUL_VIOLATOR",
]


def fy(path):
    m = re.search(r"FY(\d{4})", path)
    return m.group(1) if m else "?"


def main():
    if len(sys.argv) < 5:
        sys.exit(__doc__)
    state = sys.argv[1].strip().upper()
    cities = [c.strip().lower() for c in sys.argv[2].split(",") if c.strip()]
    out_path = sys.argv[3]
    sources = sys.argv[4:]

    matched = 0
    with open(out_path, "w", newline="") as fh:
        w = csv.writer(fh)
        w.writerow(["FISCAL_YEAR"] + KEEP)
        for src in sources:
            y = fy(src)
            wb = openpyxl.load_workbook(src, read_only=True)
            ws = wb[wb.sheetnames[0]]
            it = ws.iter_rows(values_only=True)
            header = list(next(it))
            idx = {n: header.index(n) for n in header}
            ci, si = idx["WORKSITE_CITY"], idx["WORKSITE_STATE"]
            keep_i = [idx[c] for c in KEEP]
            n = 0
            for row in it:
                n += 1
                if str(row[si] or "").strip().upper() != state:
                    continue
                city = str(row[ci] or "").strip().lower()
                if cities and not any(c in city for c in cities):
                    continue
                w.writerow([y] + [row[i] for i in keep_i])
                matched += 1
            wb.close()
            print(f"  scanned {n:,} in {Path(src).name}", file=sys.stderr)
    print(f"matched {matched} -> {out_path}", file=sys.stderr)


if __name__ == "__main__":
    main()

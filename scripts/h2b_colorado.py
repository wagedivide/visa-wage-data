import openpyxl, collections, statistics
wb=openpyxl.load_workbook("data/H-2B_FY2025.xlsx", read_only=True); ws=wb[wb.sheetnames[0]]
it=ws.iter_rows(values_only=True); next(it)
EMP,SOC,JT,CITY,STATE=17,9,7,85,86
WAGE_F,WAGE_T,PER,NWC,NWR=90,91,92,11,10
BEG,END,HRS,LODGE,DEDUCT,STATUS=14,13,67,106,107,1
RESORT={"vail","aspen","breckenridge","telluride","steamboat springs","winter park","keystone","beaver creek","snowmass","snowmass village","crested butte","copper mountain","avon","edwards","dillon","silverthorne","frisco","granby","estes park","durango","mountain village","mt crested butte","mt. crested butte","glenwood springs"}
rows=[]; empties=0; n=0
for r in it:
    if r[0] is None:
        empties+=1
        if empties>2000: break
        continue
    empties=0; n+=1
    if str(r[STATE] or "").upper()!="CO": continue
    if "certif" not in str(r[STATUS] or "").lower(): continue
    rows.append(r)
print(f"CO certified H-2B applications: {len(rows)} (of {n:,} real rows)")
def workers(r):
    for i in (NWC,NWR):
        try: return int(r[i])
        except: pass
    return 0
tot=sum(workers(r) for r in rows)
print(f"total worker positions certified: {tot:,}")
def hr(r):
    if str(r[PER] or '').lower().startswith('hour'):
        try: return float(r[WAGE_F])
        except: return None
    return None
print("\nTOP CO H-2B EMPLOYERS (by positions):")
byemp=collections.Counter()
for r in rows: byemp[str(r[EMP] or "")]+=workers(r)
for e,c in byemp.most_common(15): print(f"  {c:>4}  {e[:48]}")
print("\nTOP WORKSITE CITIES (by positions):")
bycity=collections.Counter()
for r in rows: bycity[str(r[CITY] or "").title()]+=workers(r)
for c,k in bycity.most_common(15): print(f"  {k:>4}  {c}")
print("\nTOP SOC TITLES (by positions):")
bysoc=collections.Counter()
for r in rows: bysoc[str(r[SOC] or "")]+=workers(r)
for s,k in bysoc.most_common(12): print(f"  {k:>4}  {s[:45]}")
allhr=[hr(r) for r in rows]; allhr=[x for x in allhr if x]
print(f"\nHOURLY WAGE (all CO H-2B): n={len(allhr)} min ${min(allhr):.2f} median ${statistics.median(allhr):.2f} mean ${statistics.mean(allhr):.2f} max ${max(allhr):.2f}")
res=[r for r in rows if str(r[CITY] or "").lower().strip() in RESORT]
rtot=sum(workers(r) for r in res)
rhr=[hr(r) for r in res]; rhr=[x for x in rhr if x]
print(f"\nRESORT-TOWN: {len(res)} applications, {rtot:,} positions")
if rhr: print(f"  resort hourly: median ${statistics.median(rhr):.2f}  range ${min(rhr):.2f}-${max(rhr):.2f}")
lodge=sum(1 for r in res if str(r[LODGE] or '').strip().upper() in ('Y','YES'))
print(f"  provide board/lodging: {lodge}/{len(res)}")
print("\nSAMPLE RESORT ROLES (soc | wage | #pos | employer — city):")
for r in sorted(res,key=lambda x:-workers(x))[:18]:
    w=r[WAGE_F]
    try: w=f"${float(w):.2f}/hr"
    except: pass
    print(f"  {str(r[SOC] or '')[:28]:28} {str(w):>11}  x{workers(r):<3} {str(r[EMP] or '')[:26]} — {str(r[CITY] or '').title()}")
wb.close()

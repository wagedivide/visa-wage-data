#!/usr/bin/env python3
"""Join PERM green-card filings to the PW disclosure file to recover the
prevailing-wage AMOUNT + wage LEVEL (the PERM new form omits both).

PERM.JOB_OPP_PWD_NUMBER == PW.CASE_NUMBER (the prevailing-wage determination id).

Usage:
    join_perm_pw.py "<employer regex>" <WORKSITE_STATE> \
        --perm data/PERM_FY2025.xlsx data/PERM_FY2024_newform.xlsx \
        --pw   data/PW_FY2025.xlsx data/PW_FY2024.xlsx

Prints offered-vs-floor distribution + wage-level mix for matched rows.
"""
import collections
import re
import statistics
import sys

import openpyxl


def annual(v, unit):
    try:
        v = float(v)
    except (TypeError, ValueError):
        return None
    u = (unit or "").lower()
    if u.startswith("hour"):
        return v * 2080
    if u.startswith("week"):
        return v * 52
    if u.startswith("month"):
        return v * 12
    if u.startswith("bi"):
        return v * 26
    return v


def main():
    pat = re.compile(sys.argv[1], re.I)
    state = sys.argv[2].strip().upper()
    perm_files = []
    pw_files = []
    bucket = None
    for a in sys.argv[3:]:
        if a == "--perm":
            bucket = perm_files
        elif a == "--pw":
            bucket = pw_files
        else:
            bucket.append(a)

    # 1) pull matching PERM rows
    perm = []  # dict per row
    for f in perm_files:
        wb = openpyxl.load_workbook(f, read_only=True)
        ws = wb[wb.sheetnames[0]]
        it = ws.iter_rows(values_only=True)
        h = {n: i for i, n in enumerate(next(it))}
        emp = h.get("EMP_BUSINESS_NAME", h.get("EMPLOYER_NAME"))
        pwd = h.get("JOB_OPP_PWD_NUMBER")
        wf = h.get("JOB_OPP_WAGE_FROM")
        wp = h.get("JOB_OPP_WAGE_PER")
        jt = h.get("JOB_TITLE")
        wcity = h.get("PRIMARY_WORKSITE_CITY")
        wstate = h.get("PRIMARY_WORKSITE_STATE")
        st = h.get("CASE_STATUS")
        if pwd is None:
            print(f"  (skip {f}: no JOB_OPP_PWD_NUMBER)", file=sys.stderr)
            wb.close()
            continue
        for r in it:
            if not pat.search(str(r[emp] or "")):
                continue
            if str(r[wstate] or "").upper() != state:
                continue
            perm.append({
                "pwd": str(r[pwd] or "").strip(),
                "offer": annual(r[wf], r[wp] if wp is not None else "Year"),
                "title": str(r[jt] or ""),
                "city": str(r[wcity] or ""),
                "status": str(r[st] or ""),
            })
        wb.close()
    needed = {p["pwd"] for p in perm if p["pwd"]}
    print(f"PERM {state} rows: {len(perm)}   distinct PWD ids: {len(needed)}", file=sys.stderr)

    # 2) scan PW files, collect only needed determinations
    look = {}
    for f in pw_files:
        wb = openpyxl.load_workbook(f, read_only=True)
        ws = wb[wb.sheetnames[0]]
        it = ws.iter_rows(values_only=True)
        h = {n: i for i, n in enumerate(next(it))}
        cn, wr, up, lv = (h["CASE_NUMBER"], h["PWD_WAGE_RATE"],
                          h["PWD_UNIT_OF_PAY"], h["PWD_OES_WAGE_LEVEL"])
        n = 0
        for r in it:
            n += 1
            cid = str(r[cn] or "").strip()
            if cid in needed and cid not in look:
                look[cid] = (annual(r[wr], r[up]), str(r[lv] or ""))
        wb.close()
        print(f"  scanned {n:,} PW rows in {f}", file=sys.stderr)

    # 3) join + report
    matched = [p for p in perm if p["pwd"] in look]
    print(f"\nmatched PERM↔PW: {len(matched)}/{len(perm)} "
          f"({len(matched)/len(perm)*100:.0f}% coverage)\n")
    prem = []
    levels = collections.Counter()
    for p in matched:
        pw_amt, lvl = look[p["pwd"]]
        levels[lvl or "blank"] += 1
        if p["offer"] and pw_amt and pw_amt > 1000:
            prem.append(p["offer"] / pw_amt - 1)
    print("wage-level mix (PERM green-card roles):", dict(levels))
    lvl_norm = {"Level I": "I", "Level II": "II", "Level III": "III", "Level IV": "IV"}
    below = sum(levels[k] for k in levels if lvl_norm.get(k) in ("I", "II"))
    lev = sum(levels[k] for k in levels if lvl_norm.get(k) in ("I", "II", "III", "IV"))
    if lev:
        print(f"below-median (I/II): {below}/{lev} ({below/lev*100:.0f}%)")
    if prem:
        print(f"\noffered vs prevailing floor: median {statistics.median(prem)*100:+.1f}%  "
              f"mean {statistics.mean(prem)*100:+.1f}%")
        for thr, lab in [(0.005, "at floor (<=0.5%)"), (0.05, "<=5%"), (0.10, "<=10%")]:
            print(f"  within {lab:18}: {sum(v <= thr for v in prem)}/{len(prem)} "
                  f"({sum(v <= thr for v in prem)/len(prem)*100:.0f}%)")


if __name__ == "__main__":
    main()

import openpyxl, collections, sys

def scan_lca(path, out):
    wb=openpyxl.load_workbook(path, read_only=True); ws=wb[wb.sheetnames[0]]
    it=ws.iter_rows(values_only=True); h={n:i for i,n in enumerate(next(it))}
    for r in it:
        if str(r[h["WORKSITE_STATE"]] or "").upper()!="CO": continue
        soc=str(r[h["SOC_TITLE"]] or "").lower(); jt=str(r[h["JOB_TITLE"]] or "").lower()
        naics=str(r[h["NAICS_CODE"]] or "")
        if naics.startswith("722") or "cook" in soc or "chef" in soc or "cook" in jt or "chef" in jt or "food prep" in soc:
            out.append(("LCA", str(r[h["EMPLOYER_NAME"]] or ""), str(r[h["WORKSITE_CITY"]] or ""), str(r[h["JOB_TITLE"]] or ""), r[h["WAGE_RATE_OF_PAY_FROM"]]))
    wb.close()

def scan_perm(path, out):
    wb=openpyxl.load_workbook(path, read_only=True); ws=wb[wb.sheetnames[0]]
    it=ws.iter_rows(values_only=True); h={n:i for i,n in enumerate(next(it))}
    def g(n): return h.get(n)
    for r in it:
        st=g("PRIMARY_WORKSITE_STATE")
        if st is None or str(r[st] or "").upper()!="CO": continue
        soc=str(r[g("PWD_SOC_TITLE")] or "").lower() if g("PWD_SOC_TITLE") else ""
        jt=str(r[g("JOB_TITLE")] or "").lower() if g("JOB_TITLE") else ""
        naics=str(r[g("EMP_NAICS")] or "") if g("EMP_NAICS") else ""
        if naics.startswith("722") or "cook" in soc or "chef" in soc or "cook" in jt or "chef" in jt or "food prep" in soc:
            out.append(("PERM", str(r[g("EMP_BUSINESS_NAME")] or ""), str(r[g("PRIMARY_WORKSITE_CITY")] or ""), str(r[g("JOB_TITLE")] or ""), r[g("JOB_OPP_WAGE_FROM")]))
    wb.close()

out=[]
scan_lca("data/LCA_FY2025.xlsx", out); scan_lca("data/LCA_FY2024.xlsx", out)
scan_perm("data/PERM_FY2025.xlsx", out); scan_perm("data/PERM_FY2024_newform.xlsx", out)
print(f"CO food-service visa rows found: {len(out)}")
print("by program:", dict(collections.Counter(x[0] for x in out)))
print("\ntop employers:")
for emp,c in collections.Counter(x[1] for x in out).most_common(20):
    print(f"  {c:>3}  {emp[:55]}")
print("\ntop cities:", dict(collections.Counter(x[2].title() for x in out).most_common(12)))
print("\nsample roles:")
for row in out[:20]:
    w=row[4]
    try: w=f"${float(w):,.0f}"
    except: pass
    print(f"  [{row[0]}] {row[3][:34]:34} {str(w):>12}  {row[1][:32]} — {row[2].title()}")

import openpyxl, collections, sys
ZIP=sys.argv[1] if len(sys.argv)>1 else "80027"
def scan(path, prog, empc, zipc, jtc, socc, wagec, statusc, cityc):
    wb=openpyxl.load_workbook(path, read_only=True); ws=wb[wb.sheetnames[0]]
    it=ws.iter_rows(values_only=True); h={n:i for i,n in enumerate(next(it))}
    def idx(name): return h.get(name)
    iz=idx(zipc)
    if iz is None: 
        wb.close(); return []
    ie,ij,iso,iw,ist,ic=idx(empc),idx(jtc),idx(socc),idx(wagec),idx(statusc),idx(cityc)
    out=[]; empties=0
    for r in it:
        if r[0] is None:
            empties+=1
            if empties>2000: break
            continue
        empties=0
        z=str(r[iz] or "").strip()
        if z[:5]!=ZIP: continue
        st=str(r[ist] or "") if ist is not None else ""
        out.append((prog, str(r[ie] or ""), str(r[ij] or "") if ij is not None else "",
                    str(r[iso] or "") if iso is not None else "", r[iw] if iw is not None else None,
                    st, str(r[ic] or "") if ic is not None else ""))
    wb.close(); return out

res=[]
res+=scan("data/LCA_FY2025.xlsx","H-1B","EMPLOYER_NAME","WORKSITE_POSTAL_CODE","JOB_TITLE","SOC_TITLE","WAGE_RATE_OF_PAY_FROM","CASE_STATUS","WORKSITE_CITY")
res+=scan("data/LCA_FY2024.xlsx","H-1B","EMPLOYER_NAME","WORKSITE_POSTAL_CODE","JOB_TITLE","SOC_TITLE","WAGE_RATE_OF_PAY_FROM","CASE_STATUS","WORKSITE_CITY")
res+=scan("data/PERM_FY2025.xlsx","PERM","EMP_BUSINESS_NAME","PRIMARY_WORKSITE_POSTAL_CODE","JOB_TITLE","PWD_SOC_TITLE","JOB_OPP_WAGE_FROM","CASE_STATUS","PRIMARY_WORKSITE_CITY")
res+=scan("data/PERM_FY2024_newform.xlsx","PERM","EMP_BUSINESS_NAME","PRIMARY_WORKSITE_POSTAL_CODE","JOB_TITLE","PWD_SOC_TITLE","JOB_OPP_WAGE_FROM","CASE_STATUS","PRIMARY_WORKSITE_CITY")
res+=scan("data/H-2B_FY2025.xlsx","H-2B","EMPLOYER_NAME","WORKSITE_POSTAL_CODE","JOB_TITLE","SOC_TITLE","BASIC_WAGE_RATE_FROM","CASE_STATUS","WORKSITE_CITY")

print(f"ZIP {ZIP}: {len(res)} filings across all programs")
print("by program:", dict(collections.Counter(x[0] for x in res)))
print("\nby employer:")
for emp,c in collections.Counter(x[1] for x in res).most_common(25):
    print(f"  {c:>3}  {emp[:52]}")
print("\ncities seen:", dict(collections.Counter(x[6].title() for x in res).most_common()))
print("\nall roles (prog | title | wage | status | employer):")
def money(w):
    try: return f"${float(w):,.0f}"
    except: return str(w)
for prog,emp,jt,soc,w,st,city in sorted(res, key=lambda x:(x[0],x[1])):
    print(f"  [{prog:4}] {jt[:34]:34} {money(w):>11}  {st[:12]:12} {emp[:30]}")
